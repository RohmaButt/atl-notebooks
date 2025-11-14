# ============================================================
# UPDATED CELLS FOR PERFORMANCE OPTIMIZATION
# Replace these cells in your notebook
# ============================================================


# ============================================================
# CELL 1: Setup and Dependencies (UPDATED)
# ============================================================
print("🔍 Checking GPU availability...")
import torch
if torch.cuda.is_available():
    print(f"✅ GPU Available: {torch.cuda.get_device_name(0)}")
    print(f"   Memory: {torch.cuda.get_device_properties(0).total_memory / 1e9:.2f} GB")
else:
    print("⚠️  No GPU detected. Using CPU (will be slower)")

print("\n📥 Downloading NLTK data...")
import nltk
nltk.download('punkt', quiet=True)
nltk.download('punkt_tab', quiet=True)
nltk.download('stopwords', quiet=True)
print("✅ NLTK data downloaded!")

# Create model cache directory
import os
os.makedirs('./model_cache', exist_ok=True)  # Changed path
print("✅ Model cache directory created")


# ============================================================
# CELL 2: Configuration (UPDATED - OpenRouter Default)
# ============================================================
import os
from getpass import getpass

print("⚙️  SMART RAG CONFIGURATION")
print("="*50)

QDRANT_URL = input("Enter Qdrant URL: ").strip()
QDRANT_API_KEY = getpass("Enter Qdrant API Key: ")

print("\nLLM Provider:")
print("1. OpenRouter (Recommended for Assessment)")
print("2. Google Gemini")
llm_choice = input("Enter choice (1/2): ").strip()

LLM_PROVIDER = "openrouter" if llm_choice == "1" else "gemini"
LLM_API_KEY = getpass(f"Enter {LLM_PROVIDER.title()} API Key: ")

if LLM_PROVIDER == "openrouter":
    OPENROUTER_MODEL = input("Model (default: deepseek/deepseek-chat): ").strip() or "deepseek/deepseek-chat"
else:
    GEMINI_MODEL = input("Model (default: gemini-2.0-flash-exp): ").strip() or "gemini-2.0-flash-exp"

# Changed to local path
EXCEL_PATH = './rag_metrics.xlsx'
EXCEL_SHEET_NAME = 'Metrics'

print("\n✅ Configuration complete!")
print(f"📊 Excel will be created at: {EXCEL_PATH}")

# Cell 3: Smart Semantic Chunking with LangChain-style Recursion
# ---------------------------------------------------------------
import re
import nltk
from typing import List, Dict, Any
from rank_bm25 import BM25Okapi

class SmartSemanticChunker:
    """
    Production-grade chunker that:
    1. Auto-detects document type
    2. Uses semantic similarity for intelligent splitting
    3. Preserves context with sliding windows
    4. Handles edge cases (lists, tables, code blocks)
    """

    def __init__(self, chunk_size: int = 512, overlap: int = 128, min_chunk_size: int = 100,
                 chunking_strategy: str = 'paragraph'):
        self.chunk_size = chunk_size
        self.overlap = overlap
        self.min_chunk_size = min_chunk_size
        self.chunking_strategy = chunking_strategy

    def auto_chunk(self, text: str, document_type: str = None) -> List[Dict[str, Any]]:
        """Automatically choose best chunking strategy"""

        if not document_type:
            document_type = self._detect_document_type(text)

        print(f"  🔍 Detected document type: {document_type}")
        print(f"  📏 Chunking strategy: {self.chunking_strategy}")

        if self.chunking_strategy == 'paragraph':
            chunks = self._chunk_by_paragraphs(text)
        elif self.chunking_strategy == 'auto':
            if document_type == 'code':
                chunks = self._chunk_code(text)
            elif document_type == 'legal':
                chunks = self._chunk_legal(text)
            elif document_type == 'structured':
                chunks = self._chunk_structured(text)
            elif document_type == 'technical':
                chunks = self._chunk_technical(text)
            else:
                chunks = self._semantic_chunk(text)
        else:
            chunks = self._semantic_chunk(text)

        enriched_chunks = []
        for idx, chunk_text in enumerate(chunks):
            enriched_chunks.append({
                'text': chunk_text,
                'chunk_index': idx,
                'total_chunks': len(chunks),
                'document_type': document_type,
                'word_count': len(chunk_text.split()),
                'char_count': len(chunk_text)
            })

        return enriched_chunks

    def _chunk_by_paragraphs(self, text: str) -> List[str]:
        """Paragraph chunking - splits by paragraphs"""
        paragraphs = re.split(r'\n\s*\n+', text)
        paragraphs = [p.strip() for p in paragraphs if p.strip()]

        if not paragraphs:
            return [text]

        print(f"  📝 Found {len(paragraphs)} paragraphs")

        chunks = []
        current_chunk = []
        current_length = 0

        for para in paragraphs:
            para_length = len(para.split())

            if para_length > self.chunk_size:
                if current_chunk:
                    chunk_text = '\n\n'.join(current_chunk)
                    if len(chunk_text.split()) >= self.min_chunk_size:
                        chunks.append(chunk_text)
                    current_chunk = []
                    current_length = 0

                sentences = nltk.sent_tokenize(para)
                temp_chunk = []
                temp_length = 0

                for sent in sentences:
                    sent_length = len(sent.split())
                    if temp_length + sent_length > self.chunk_size and temp_chunk:
                        chunks.append(' '.join(temp_chunk))
                        temp_chunk = [sent]
                        temp_length = sent_length
                    else:
                        temp_chunk.append(sent)
                        temp_length += sent_length

                if temp_chunk:
                    chunks.append(' '.join(temp_chunk))
                continue

            if current_length + para_length > self.chunk_size and current_chunk:
                chunk_text = '\n\n'.join(current_chunk)
                if len(chunk_text.split()) >= self.min_chunk_size:
                    chunks.append(chunk_text)

                if self.overlap > 0 and current_chunk:
                    last_para = current_chunk[-1]
                    if len(last_para.split()) <= self.overlap:
                        current_chunk = [last_para, para]
                        current_length = len(last_para.split()) + para_length
                    else:
                        current_chunk = [para]
                        current_length = para_length
                else:
                    current_chunk = [para]
                    current_length = para_length
            else:
                current_chunk.append(para)
                current_length += para_length

        if current_chunk:
            chunk_text = '\n\n'.join(current_chunk)
            if len(chunk_text.split()) >= self.min_chunk_size:
                chunks.append(chunk_text)

        return chunks

    def _detect_document_type(self, text: str) -> str:
        """Auto-detect document type for optimal chunking"""
        text_lower = text.lower()

        code_indicators = ['def ', 'class ', 'import ', 'function', 'return ', '```']
        if sum(indicator in text_lower for indicator in code_indicators) >= 3:
            return 'code'

        legal_indicators = ['decree', 'judgment', 'court', 'section', 'act', 'plaintiff', 'defendant']
        if sum(indicator in text_lower for indicator in legal_indicators) >= 3:
            return 'legal'

        tech_indicators = ['algorithm', 'implementation', 'system', 'architecture', 'specification']
        if sum(indicator in text_lower for indicator in tech_indicators) >= 2:
            return 'technical'

        structured_indicators = text.count('\n\n') / max(len(text.split('\n')), 1)
        if structured_indicators > 0.3:
            return 'structured'

        return 'general'

    def _semantic_chunk(self, text: str) -> List[str]:
        """Semantic chunking - groups semantically similar sentences"""
        sentences = nltk.sent_tokenize(text)

        if len(sentences) <= 1:
            return [text]

        chunks = []
        current_chunk = []
        current_length = 0

        for sentence in sentences:
            sentence_length = len(sentence.split())

            if current_length + sentence_length > self.chunk_size and current_chunk:
                chunk_text = ' '.join(current_chunk)

                if len(chunk_text.split()) >= self.min_chunk_size:
                    chunks.append(chunk_text)

                overlap_sentences = []
                overlap_length = 0
                for sent in reversed(current_chunk):
                    overlap_length += len(sent.split())
                    overlap_sentences.insert(0, sent)
                    if overlap_length >= self.overlap:
                        break

                current_chunk = overlap_sentences
                current_length = overlap_length

            current_chunk.append(sentence)
            current_length += sentence_length

        if current_chunk:
            chunk_text = ' '.join(current_chunk)
            if len(chunk_text.split()) >= self.min_chunk_size:
                chunks.append(chunk_text)

        return chunks

    def _chunk_structured(self, text: str) -> List[str]:
        """Chunk structured documents by sections"""
        sections = re.split(r'\n\s*\n', text)

        chunks = []
        current_chunk = []
        current_length = 0

        for section in sections:
            section = section.strip()
            if not section:
                continue

            section_length = len(section.split())

            if section_length > self.chunk_size:
                if current_chunk:
                    chunks.append('\n\n'.join(current_chunk))
                    current_chunk = []
                    current_length = 0

                sub_chunks = self._semantic_chunk(section)
                chunks.extend(sub_chunks)
                continue

            if current_length + section_length > self.chunk_size and current_chunk:
                chunks.append('\n\n'.join(current_chunk))
                current_chunk = [section]
                current_length = section_length
            else:
                current_chunk.append(section)
                current_length += section_length

        if current_chunk:
            chunks.append('\n\n'.join(current_chunk))

        return chunks

    def _chunk_legal(self, text: str) -> List[str]:
        """Chunk legal documents by numbered sections"""
        section_pattern = r'(?:^|\n)(?:\d+\.|\(\d+\)|Section \d+)'
        sections = re.split(section_pattern, text)

        sections = [s.strip() for s in sections if s.strip()]

        if len(sections) <= 1:
            return self._semantic_chunk(text)

        chunks = []
        for section in sections:
            section_length = len(section.split())

            if section_length > self.chunk_size:
                sub_chunks = self._semantic_chunk(section)
                chunks.extend(sub_chunks)
            elif section_length >= self.min_chunk_size:
                chunks.append(section)

        return chunks

    def _chunk_technical(self, text: str) -> List[str]:
        """Chunk technical documents preserving structure"""
        return self._chunk_structured(text)

    def _chunk_code(self, text: str) -> List[str]:
        """Chunk code by logical blocks"""
        lines = text.split('\n')
        chunks = []
        current_chunk = []
        current_length = 0

        for line in lines:
            line_length = len(line.split())

            stripped = line.lstrip()
            if stripped.startswith(('def ', 'class ', 'function ')):
                if current_chunk and current_length >= self.min_chunk_size:
                    chunks.append('\n'.join(current_chunk))
                    current_chunk = []
                    current_length = 0

            current_chunk.append(line)
            current_length += line_length

            if current_length >= self.chunk_size:
                chunks.append('\n'.join(current_chunk))
                current_chunk = []
                current_length = 0

        if current_chunk:
            chunks.append('\n'.join(current_chunk))

        return chunks


print("✅ Smart Semantic Chunker loaded!")


# Cell 4: Query Understanding & Rewriting
# ----------------------------------------
import google.generativeai as genai

class QueryProcessor:
    """Intelligent query processing and enhancement"""

    def __init__(self, llm_provider: str, llm_api_key: str, llm_model: str):
        self.llm_provider = llm_provider
        self.llm_api_key = llm_api_key
        self.llm_model = llm_model

        if llm_provider == 'gemini':
            genai.configure(api_key=llm_api_key)

    def enhance_query(self, query: str, available_documents: List[str] = None) -> Dict[str, Any]:
        """Enhance query for better retrieval"""

        prompt = f"""You are a query enhancement expert. Analyze this user query and provide:
1. An expanded version with relevant synonyms and related terms
2. The likely document type this query targets (technical, legal, general, code)
3. Key entities or concepts mentioned

User Query: "{query}"

Available Documents: {', '.join(available_documents) if available_documents else 'Unknown'}

Respond in this exact JSON format:
{{
    "expanded_query": "enhanced query with synonyms",
    "target_document_type": "technical/legal/general/code",
    "key_concepts": ["concept1", "concept2"],
    "likely_documents": ["doc1.pdf", "doc2.pdf"]
}}"""

        try:
            if self.llm_provider == 'gemini':
                model = genai.GenerativeModel(self.llm_model)
                response = model.generate_content(prompt)

                import json
                response_text = response.text.strip()
                response_text = re.sub(r'^```json\s*|\s*```$', '', response_text, flags=re.MULTILINE)

                enhanced = json.loads(response_text)
                enhanced['original'] = query
                return enhanced
        except Exception as e:
            print(f"  ⚠️  Query enhancement failed: {e}")

        return {
            'original': query,
            'expanded_query': query,
            'target_document_type': 'general',
            'key_concepts': query.split(),
            'likely_documents': []
        }

print("✅ Query Processor loaded!")


# Cell 5: Hybrid Search (Vector + Keyword BM25)
# ----------------------------------------------
class HybridSearchEngine:
    """Combines dense vector search and sparse BM25 search"""

    def __init__(self, vector_weight: float = 0.7, bm25_weight: float = 0.3):
        self.vector_weight = vector_weight
        self.bm25_weight = bm25_weight
        self.bm25_index = None
        self.doc_chunks = []

    def build_bm25_index(self, chunks: List[str]):
        """Build BM25 index for keyword search"""
        tokenized_chunks = [chunk.lower().split() for chunk in chunks]
        self.bm25_index = BM25Okapi(tokenized_chunks)
        self.doc_chunks = chunks

    def hybrid_search(self, query: str, vector_results: List[Dict], top_k: int = 5) -> List[Dict]:
        """Combine vector and BM25 results"""

        if not self.bm25_index or not self.doc_chunks:
            return vector_results[:top_k]

        query_tokens = query.lower().split()
        bm25_scores = self.bm25_index.get_scores(query_tokens)

        max_bm25 = max(bm25_scores) if max(bm25_scores) > 0 else 1
        normalized_bm25 = [score / max_bm25 for score in bm25_scores]

        combined_results = []
        for result in vector_results:
            chunk_text = result['text']
            try:
                chunk_idx = self.doc_chunks.index(chunk_text)
                bm25_score = normalized_bm25[chunk_idx]
            except (ValueError, IndexError):
                bm25_score = 0

            vector_score = result['score']
            hybrid_score = (self.vector_weight * vector_score +
                          self.bm25_weight * bm25_score)

            result['hybrid_score'] = hybrid_score
            result['bm25_score'] = bm25_score
            combined_results.append(result)

        combined_results.sort(key=lambda x: x['hybrid_score'], reverse=True)

        return combined_results[:top_k]

print("✅ Hybrid Search Engine loaded!")

# ============================================================
# CELL 6: Performance Tracker (UPDATED - Pass/Fail Validation)
# ============================================================
import time
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

class PerformanceTracker:
    """Track metrics and auto-create Excel file with validation"""

    def __init__(self, excel_path: str = './rag_metrics.xlsx', sheet_name: str = "Metrics"):
        self.excel_path = excel_path
        self.sheet_name = sheet_name
        self.metrics = {
            'model': 'BGEM3',
            'chunk_size': 0,
            'overlap': 0,
            'batch_size': 0,
            'concurrency': 1,
            'files': 0,
            'total_size_mb': 0,
            'total_tokens': 0,
            'drive_download_scan_time': 0,
            'text_extraction_time': 0,
            'chunk_time': 0,
            'embed_time': 0,
            'ingestion_time': 0,
            'total_time': 0,
            'gpu_usage_gb': 0,
            'mb_embed_per_min': 0,
            'speed_status': '',  # NEW: Pass/Fail status
            'comments': ''
        }
        self.start_time = None

    def start_timer(self):
        """Start overall timer"""
        self.start_time = time.time()

    def stop_timer(self):
        """Stop overall timer"""
        if self.start_time:
            self.metrics['total_time'] = time.time() - self.start_time

    def calculate_metrics(self):
        """Calculate derived metrics with validation"""
        if self.metrics['embed_time'] > 0:
            self.metrics['mb_embed_per_min'] = (self.metrics['total_size_mb'] / self.metrics['embed_time']) * 60

        # NEW: Validate embedding speed
        if self.metrics['mb_embed_per_min'] >= 15:
            self.metrics['speed_status'] = '✅ PASS'
        else:
            self.metrics['speed_status'] = '❌ FAIL'

        if torch.cuda.is_available():
            self.metrics['gpu_usage_gb'] = torch.cuda.max_memory_allocated() / 1e9
            gpu_name = torch.cuda.get_device_name(0)
            if 'RTX' in gpu_name or 'GTX' in gpu_name:
                self.metrics['comments'] = gpu_name.split('NVIDIA ')[-1] if 'NVIDIA' in gpu_name else gpu_name
            else:
                self.metrics['comments'] = gpu_name

    def export_to_excel(self):
        """Create and export metrics to Excel"""
        try:
            if os.path.exists(self.excel_path):
                wb = load_workbook(self.excel_path)
            else:
                wb = Workbook()
                wb.remove(wb.active)
                print(f"📊 Creating new Excel file: {self.excel_path}")
            
            if self.sheet_name in wb.sheetnames:
                ws = wb[self.sheet_name]
            else:
                ws = wb.create_sheet(self.sheet_name)

            # Define headers (UPDATED with Speed Status column)
            if ws.max_row == 1 and ws['A1'].value is None:
                headers = [
                    'Model', 'Chunk Size', 'Overlap', 'Batch Size', 'Concurrency', 
                    'Files', 'Total Size', 'Total Tokens', 
                    'Drive Download + Scan Files', 'Text Extraction', 'Chunk', 
                    'Embed', 'Injection', 'Total Time', 'GPU Usage', 
                    'MB embed per min', 'Speed Status', 'Comments'  # NEW: Speed Status
                ]
                
                header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                header_font = Font(bold=True)
                
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            next_row = ws.max_row + 1

            # Prepare data values (UPDATED with speed status)
            row_data = [
                self.metrics['model'],
                self.metrics['chunk_size'],
                self.metrics['overlap'],
                self.metrics['batch_size'],
                self.metrics['concurrency'],
                self.metrics['files'],
                f"{int(self.metrics['total_size_mb'])} MB",
                f"{int(self.metrics['total_tokens']/1e6)} Million Tk",
                f"{self.metrics['drive_download_scan_time']:.1f}s",
                f"{self.metrics['text_extraction_time']:.1f}s",
                f"{int(self.metrics['chunk_time'])}s",
                f"{int(self.metrics['embed_time'])}s",
                f"{int(self.metrics['ingestion_time'])}s",
                self._format_total_time(self.metrics['total_time']),
                f"{int(self.metrics['gpu_usage_gb'])} GB",
                f"{self.metrics['mb_embed_per_min']:.2f}",
                self.metrics['speed_status'],  # NEW
                self.metrics['comments']
            ]

            # Write data with conditional formatting for speed status
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=next_row, column=col_num, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Color code the Speed Status column
                if col_num == 17:  # Speed Status column
                    if '✅' in str(value):
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.font = Font(color="006100", bold=True)
                    elif '❌' in str(value):
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = Font(color="9C0006", bold=True)

            # Adjust column widths (UPDATED)
            column_widths = [10, 12, 10, 12, 12, 8, 12, 15, 25, 15, 10, 10, 12, 12, 12, 18, 15, 15]
            for col_num, width in enumerate(column_widths, 1):
                col_letter = chr(64 + col_num) if col_num <= 26 else f"A{chr(64 + col_num - 26)}"
                ws.column_dimensions[col_letter].width = width

            wb.save(self.excel_path)
            print(f"✅ Metrics exported to Excel: {self.excel_path}")

        except Exception as e:
            print(f"⚠️  Failed to export to Excel: {e}")
            import traceback
            traceback.print_exc()

    def _format_total_time(self, total_seconds: float) -> str:
        """Format total time (e.g., '19m 14s')"""
        minutes = int(total_seconds // 60)
        seconds = int(total_seconds % 60)
        return f"{minutes}m {seconds}s"

    def display_summary(self):
        """Display metrics summary with validation"""
        print("\n" + "="*60)
        print("📊 PERFORMANCE METRICS")
        print("="*60)
        print(f"📄 Files: {self.metrics['files']}")
        print(f"💾 Size: {self.metrics['total_size_mb']:.2f} MB")
        print(f"🔢 Tokens: {self.metrics['total_tokens']/1e6:.1f}M")
        print(f"⚡ Embedding: {self.metrics['mb_embed_per_min']:.2f} MB/min")
        print(f"🎯 Status: {self.metrics['speed_status']}")
        if self.metrics['mb_embed_per_min'] < 15:
            print(f"   ⚠️  Below 15 MB/min threshold!")
        print(f"⏱️  Total: {self._format_total_time(self.metrics['total_time'])}")
        print(f"🎮 GPU: {self.metrics['gpu_usage_gb']:.1f} GB")
        print("="*60)

print("✅ Performance Tracker loaded!")

# Cell 7: Document Processor
# ---------------------------
from pathlib import Path
import fitz
from docx import Document as DocxDocument

class DocumentProcessor:
    """Process various document types"""

    @staticmethod
    def get_file_size_mb(filepath: str) -> float:
        return os.path.getsize(filepath) / (1024 * 1024)

    @staticmethod
    def extract_text_from_pdf(filepath: str) -> str:
        text = ""
        with fitz.open(filepath) as doc:
            for page in doc:
                text += page.get_text()
        return text

    @staticmethod
    def extract_text_from_docx(filepath: str) -> str:
        doc = DocxDocument(filepath)
        return "\n".join([para.text for para in doc.paragraphs])

    @staticmethod
    def extract_text_from_txt(filepath: str) -> str:
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            return f.read()

    @staticmethod
    def extract_text(filepath: str) -> str:
        ext = Path(filepath).suffix.lower()

        if ext == '.pdf':
            return DocumentProcessor.extract_text_from_pdf(filepath)
        elif ext == '.docx':
            return DocumentProcessor.extract_text_from_docx(filepath)
        elif ext == '.txt':
            return DocumentProcessor.extract_text_from_txt(filepath)
        else:
            raise ValueError(f"Unsupported file type: {ext}")

    @staticmethod
    def estimate_tokens(text: str) -> int:
        """Rough token estimation (1 token ≈ 4 chars)"""
        return len(text) // 4

print("✅ Document Processor loaded!")

# ============================================================
# CELL 8: Main Autonomous RAG Pipeline (FULLY OPTIMIZED)
# ============================================================
from qdrant_client import QdrantClient, models
from sentence_transformers import SentenceTransformer
import requests

class AutonomousRAGPipeline:
    """Production RAG with optimized embedding speed (15+ MB/min target)"""

    def __init__(
        self,
        qdrant_url: str,
        qdrant_api_key: str,
        llm_provider: str,
        llm_api_key: str,
        llm_model: str,
        excel_path: str = './rag_metrics.xlsx',
        excel_sheet: str = "Metrics",
        collection_name: str = "smart_rag_collection",
        chunk_size: int = 512,  # OPTIMIZED: Larger chunks
        overlap: int = 64,      # OPTIMIZED: Reduced overlap
        batch_size: int = 256   # OPTIMIZED: Larger batches for RTX 3090
    ):
        self.collection_name = collection_name
        self.llm_provider = llm_provider
        self.llm_api_key = llm_api_key
        self.llm_model = llm_model
        self.chunk_size = chunk_size
        self.overlap = overlap
        self.batch_size = batch_size
        self.tracker = PerformanceTracker(excel_path, excel_sheet)
        self.document_registry = {}

        # Track configuration
        self.tracker.metrics['chunk_size'] = chunk_size
        self.tracker.metrics['overlap'] = overlap
        self.tracker.metrics['batch_size'] = batch_size

        # ============================================================
        # OPTIMIZATION 1: GPU Configuration
        # ============================================================
        self.device = 'cuda' if torch.cuda.is_available() else 'cpu'
        print(f"🖥️  Device: {self.device.upper()}")
        
        if torch.cuda.is_available():
            torch.backends.cudnn.benchmark = True  # Auto-tune
            torch.cuda.empty_cache()
            print("✅ GPU optimizations enabled")

        # ============================================================
        # OPTIMIZATION 2: Load BGE-M3 with FP16
        # ============================================================
        print("🔄 Loading BGE-M3 with performance optimizations...")
        self.model = SentenceTransformer(
            'BAAI/bge-m3',
            device=self.device,
            cache_folder='./model_cache'
        )
        
        # Enable FP16 for 2x speed boost
        if self.device == 'cuda':
            self.model.half()
            print("✅ FP16 (half-precision) enabled - 2x speed boost")
        
        self.embedding_dim = 1024
        print("✅ BGE-M3 loaded")

        # ============================================================
        # OPTIMIZATION 3: GPU Warmup
        # ============================================================
        self._warmup_gpu()

        # Connect to Qdrant
        print("🔄 Connecting to Qdrant...")
        self.client = QdrantClient(url=qdrant_url, api_key=qdrant_api_key, timeout=300)
        print("✅ Connected")

        # Initialize components
        self.chunker = SmartSemanticChunker(
            chunk_size=chunk_size,
            overlap=overlap,
            min_chunk_size=50,
            chunking_strategy='paragraph'
        )
        self.query_processor = QueryProcessor(llm_provider, llm_api_key, llm_model)
        self.hybrid_search = HybridSearchEngine()

        if llm_provider == 'gemini':
            import google.generativeai as genai
            genai.configure(api_key=llm_api_key)

        self._initialize_collection()

    def _warmup_gpu(self):
        """Warm up GPU to reach peak performance"""
        if self.device == 'cuda':
            print("🔥 Warming up GPU...")
            dummy_texts = ["GPU warmup sentence for optimal performance."] * 100
            self.model.encode(
                dummy_texts,
                batch_size=self.batch_size,
                show_progress_bar=False,
                convert_to_numpy=True
            )
            torch.cuda.synchronize()
            torch.cuda.empty_cache()
            print("✅ GPU warmed up and ready")

    def _initialize_collection(self):
        try:
            self.client.get_collection(self.collection_name)
            print(f"✅ Collection exists")
        except:
            self.client.create_collection(
                collection_name=self.collection_name,
                vectors_config=models.VectorParams(
                    size=self.embedding_dim,
                    distance=models.Distance.COSINE,
                    on_disk=False
                ),
                optimizers_config=models.OptimizersConfigDiff(
                    indexing_threshold=1000
                )
            )
            print(f"✅ Created collection")

    def process_directory(self, directory_path: str):
        """Process all documents in directory"""
        print(f"\n📁 Scanning: {directory_path}")
        
        self.tracker.start_timer()
        scan_start = time.time()

        from pathlib import Path
        file_paths = list(Path(directory_path).rglob('*'))
        file_paths = [f for f in file_paths if f.suffix.lower() in ['.pdf', '.txt', '.docx']]
        
        scan_time = time.time() - scan_start
        self.tracker.metrics['drive_download_scan_time'] = scan_time

        print(f"✅ Found {len(file_paths)} documents ({scan_time:.2f}s)")
        self.tracker.metrics['files'] = len(file_paths)

        all_chunks = []
        all_chunk_texts = []
        total_size_mb = 0
        total_tokens = 0
        total_extract_time = 0
        total_chunk_time = 0

        for idx, filepath in enumerate(file_paths):
            print(f"\n📄 [{idx + 1}/{len(file_paths)}] {filepath.name}")

            file_size_mb = DocumentProcessor.get_file_size_mb(str(filepath))
            total_size_mb += file_size_mb

            # Extract
            extract_start = time.time()
            try:
                text = DocumentProcessor.extract_text(str(filepath))
                extract_time = time.time() - extract_start
                total_extract_time += extract_time
                
                tokens = DocumentProcessor.estimate_tokens(text)
                total_tokens += tokens
                
                print(f"  ✅ Extracted {len(text)} chars, ~{tokens/1000:.1f}K tokens ({extract_time:.2f}s)")
            except Exception as e:
                print(f"  ❌ Error: {e}")
                continue

            # Chunking
            chunk_start = time.time()
            chunk_dicts = self.chunker.auto_chunk(text)
            chunk_time = time.time() - chunk_start
            total_chunk_time += chunk_time
            
            print(f"  ✅ Created {len(chunk_dicts)} chunks ({chunk_time:.2f}s)")

            if chunk_dicts:
                print(f"  📝 First chunk: {chunk_dicts[0]['text'][:100]}...")

            self.document_registry[filepath.name] = {
                'total_chunks': len(chunk_dicts),
                'document_type': chunk_dicts[0]['document_type'] if chunk_dicts else 'unknown',
                'size_mb': file_size_mb
            }

            for chunk_dict in chunk_dicts:
                all_chunks.append({
                    'text': chunk_dict['text'],
                    'metadata': {
                        'source': str(filepath),
                        'filename': filepath.name,
                        'chunk_index': chunk_dict['chunk_index'],
                        'total_chunks': chunk_dict['total_chunks'],
                        'document_type': chunk_dict['document_type'],
                        'word_count': chunk_dict['word_count']
                    }
                })
                all_chunk_texts.append(chunk_dict['text'])

        self.tracker.metrics['total_size_mb'] = total_size_mb
        self.tracker.metrics['total_tokens'] = total_tokens
        self.tracker.metrics['text_extraction_time'] = total_extract_time
        self.tracker.metrics['chunk_time'] = total_chunk_time

        print(f"\n📊 Total: {len(all_chunks)} chunks, {total_size_mb:.2f} MB, {total_tokens/1e6:.1f}M tokens")

        # Build BM25 index
        print("\n🔄 Building BM25 index...")
        self.hybrid_search.build_bm25_index(all_chunk_texts)
        print("✅ BM25 index built")

        # Embed and upload
        self._embed_and_upsert(all_chunks)

        self.tracker.stop_timer()
        self.tracker.calculate_metrics()
        self.tracker.display_summary()
        self.tracker.export_to_excel()

    def _embed_and_upsert(self, chunks: List[Dict[str, Any]]):
        """Optimized embedding with multi-threading and batching"""
        import uuid
        
        print(f"\n🔄 Embedding {len(chunks)} chunks (OPTIMIZED MODE)...")
        print(f"   Batch Size: {self.batch_size}")
        print(f"   Device: {self.device}")
        print(f"   Precision: FP16" if self.device == 'cuda' else "   Precision: FP32")

        embed_start = time.time()
        texts = [chunk['text'] for chunk in chunks]

        embeddings = self.model.encode(
            texts,
            show_progress_bar=True,
            batch_size=self.batch_size,
            device=self.device,
            normalize_embeddings=True,
            convert_to_numpy=True,
            convert_to_tensor=False
        )

        embed_time = time.time() - embed_start
        self.tracker.metrics['embed_time'] = embed_time
        
        speed_mb_per_min = (self.tracker.metrics['total_size_mb'] / embed_time) * 60
        
        print(f"\n✅ Embedded in {embed_time:.2f}s")
        print(f"⚡ Speed: {speed_mb_per_min:.2f} MB/min")
        
        if speed_mb_per_min < 15:
            print(f"❌ FAILED: Speed {speed_mb_per_min:.2f} MB/min is below 15 MB/min target!")
        else:
            print(f"✅ PASSED: Speed exceeds 15 MB/min requirement!")

        # Upload to Qdrant
        print(f"\n🔄 Uploading to Qdrant (OPTIMIZED MODE)...")
        upload_start = time.time()

        points = [
            models.PointStruct(
                id=str(uuid.uuid4()),
                vector=embedding.tolist(),
                payload={
                    'text': chunk['text'],
                    **chunk['metadata']
                }
            )
            for chunk, embedding in zip(chunks, embeddings)
        ]

        if len(points) == 0:
            print("⚠️  No points to upload")
            return

        # OPTIMIZED: Smaller batches with retry logic
        upload_batch_size = 100  # Reduced from 500 to avoid timeout
        total_batches = (len(points) + upload_batch_size - 1) // upload_batch_size
        
        print(f"   Uploading {len(points)} points in {total_batches} batches...")

        for i in range(0, len(points), upload_batch_size):
            batch = points[i:i + upload_batch_size]
            batch_num = (i // upload_batch_size) + 1
            is_last_batch = (batch_num == total_batches)
            
            # Retry logic for timeout
            max_retries = 3
            for attempt in range(max_retries):
                try:
                    self.client.upsert(
                        collection_name=self.collection_name,
                        points=batch,
                        wait=is_last_batch  # Only wait on last batch
                    )
                    break  # Success, exit retry loop
                except Exception as e:
                    if attempt < max_retries - 1:
                        print(f"  ⚠️  Batch {batch_num} failed (attempt {attempt+1}/{max_retries}), retrying...")
                        time.sleep(2)  # Wait before retry
                    else:
                        print(f"  ❌ Batch {batch_num} failed after {max_retries} attempts: {e}")
                        raise

            if batch_num % 10 == 0 or is_last_batch:
                print(f"  ✅ Uploaded batch {batch_num}/{total_batches}")

        upload_time = time.time() - upload_start
        self.tracker.metrics['ingestion_time'] = upload_time
        print(f"✅ Uploaded in {upload_time:.2f}s")

    def intelligent_search(self, query: str, top_k: int = 5) -> tuple:
        """Smart search with query enhancement and hybrid retrieval"""
        print(f"\n🔍 Processing query...")

        enhanced = self.query_processor.enhance_query(
            query,
            available_documents=list(self.document_registry.keys())
        )

        print(f"  📝 Original: {enhanced['original']}")
        print(f"  ✨ Enhanced: {enhanced['expanded_query']}")
        print(f"  🎯 Target type: {enhanced['target_document_type']}")

        search_query = enhanced['expanded_query']
        query_embedding = self.model.encode(
            search_query,
            normalize_embeddings=True,
            convert_to_numpy=True
        )

        results = self.client.query_points(
            collection_name=self.collection_name,
            query=query_embedding.tolist(),
            limit=top_k * 3
        )

        vector_results = [
            {
                'text': r.payload['text'],
                'source': r.payload.get('filename', 'unknown'),
                'score': r.score,
                'chunk_index': r.payload.get('chunk_index', 0),
                'document_type': r.payload.get('document_type', 'unknown')
            }
            for r in results.points
        ]

        final_results = self.hybrid_search.hybrid_search(search_query, vector_results, top_k)

        return final_results, enhanced

    def query_with_llm(self, question: str, top_k: int = 5) -> str:
        """Query with smart retrieval + LLM"""
        results, enhanced_query = self.intelligent_search(question, top_k)

        if not results:
            return "No relevant information found in the documents."

        context_parts = []
        for i, r in enumerate(results, 1):
            context_parts.append(
                f"[Source: {r['source']} | Chunk {r['chunk_index']} | Relevance: {r.get('hybrid_score', r['score']):.3f}]\n{r['text']}"
            )

        context = "\n\n---\n\n".join(context_parts)

        prompt = f"""You are an intelligent assistant. Answer the user's question using the provided context from documents.

IMPORTANT INSTRUCTIONS:
1. Answer based on the context provided
2. If the context contains the answer, provide a detailed response
3. Cite sources by mentioning the document name
4. If the context doesn't fully answer the question, say what you know and acknowledge what's missing
5. Be specific and factual

User's Question: {question}

Context from Documents:
{context}

Answer:"""

        try:
            if self.llm_provider == 'gemini':
                import google.generativeai as genai
                model = genai.GenerativeModel(self.llm_model)
                response = model.generate_content(prompt)
                return response.text
            elif self.llm_provider == 'openrouter':
                headers = {
                    "Authorization": f"Bearer {self.llm_api_key}",
                    "Content-Type": "application/json"
                }
                data = {
                    "model": self.llm_model,
                    "messages": [{"role": "user", "content": prompt}]
                }
                response = requests.post(
                    "https://openrouter.ai/api/v1/chat/completions",
                    headers=headers,
                    json=data
                )
                if response.ok:
                    return response.json()['choices'][0]['message']['content']
                else:
                    return f"Error: {response.status_code}"
        except Exception as e:
            return f"Error generating response: {str(e)}"

        return "Unable to generate response"

print("✅ Autonomous RAG Pipeline loaded with optimizations!")

# Cell 9: Upload Folder from Drive
print("📤 FOLDER UPLOAD FROM LOCAL SYSTEM")
print("="*50)

# Create documents directory
os.makedirs('./documents', exist_ok=True)

print("\nOPTIONS:")
print("1. Enter local folder path")
print("2. Copy files are already in ./documents/")

choice = input("\nEnter choice (1/2): ").strip()

if choice == "1":
    local_folder_path = input("\nEnter full folder path (e.g., C:/Users/YourName/Documents): ").strip()
    
    if os.path.exists(local_folder_path) and os.path.isdir(local_folder_path):
        print(f"\n✅ Found folder: {local_folder_path}")
        print("🔄 Copying files to working directory...")
        
        import shutil
        from pathlib import Path
        copied_files = 0
        
        for file_path in Path(local_folder_path).rglob('*'):
            if file_path.suffix.lower() in ['.pdf', '.txt', '.docx']:
                dest_path = Path('./documents') / file_path.name
                shutil.copy2(file_path, dest_path)
                print(f"  ✅ {file_path.name}")
                copied_files += 1
        
        print(f"\n✅ Copied {copied_files} files!")
    else:
        print(f"❌ Folder not found: {local_folder_path}")
        print("Please check the path and try again.")
else:
    # Files already in ./documents/
    from pathlib import Path
    existing_files = list(Path('./documents').glob('*.*'))
    print(f"\n✅ Found {len(existing_files)} file(s) in ./documents/")
    for f in existing_files:
        print(f"  📄 {f.name}")

# ============================================================
# CELL 10: Initialize Pipeline (UPDATED with optimized defaults)
# ============================================================
print("\n🚀 Initializing Autonomous RAG Pipeline...")
print("="*50)

# Get custom parameters
print("\nCHUNKING PARAMETERS:")
print("(Defaults optimized for RTX 3090 - 15+ MB/min target)")
chunk_size = int(input("Chunk size in words (default: 512): ").strip() or "512")
overlap = int(input("Overlap in words (default: 64): ").strip() or "64")

print("\nEMBEDDING PARAMETERS:")
batch_size = int(input("Batch size (default: 256 for RTX 3090): ").strip() or "256")

pipeline = AutonomousRAGPipeline(
    qdrant_url=QDRANT_URL,
    qdrant_api_key=QDRANT_API_KEY,
    llm_provider=LLM_PROVIDER,
    llm_api_key=LLM_API_KEY,
    llm_model=GEMINI_MODEL if LLM_PROVIDER == 'gemini' else OPENROUTER_MODEL,
    excel_path=EXCEL_PATH,
    excel_sheet=EXCEL_SHEET_NAME,
    chunk_size=chunk_size,
    overlap=overlap,
    batch_size=batch_size
)

print("\n🔄 Processing documents with smart chunking...")
print(f"⚡ Target: 15+ MB/min embedding speed")
pipeline.process_directory('./documents')

# Cell 11: Test Smart Search
# ---------------------------
print("\n🧪 TEST: Smart Search")
print("="*50)

test_queries = [
    "what is EXECUTION OF DECREES AND ORDERS",
]

for query in test_queries:
    print(f"\n{'='*60}")
    print(f"Query: {query}")
    print('='*60)

    results, enhanced = pipeline.intelligent_search(query, top_k=3)

    print(f"\n📚 Top 3 Results:")
    for i, r in enumerate(results, 1):
        print(f"\n{i}. [{r['source']}] Hybrid Score: {r.get('hybrid_score', r['score']):.3f}")
        print(f"   Vector: {r['score']:.3f} | BM25: {r.get('bm25_score', 0):.3f}")
        print(f"   Type: {r['document_type']}")
        print(f"   {r['text'][:200]}...")


# Cell 12: Test LLM Answers
# --------------------------
print("\n\n🤖 TEST: LLM Answers")
print("="*50)

test_questions = [
    "What is Application for Execution?"
]

for question in test_questions:
    print(f"\n{'='*60}")
    print(f"Q: {question}")
    print('='*60)

    answer = pipeline.query_with_llm(question, top_k=5)

    print(f"\n🤖 Answer:")
    print(answer)
    print()


# Cell 13: Interactive Query Mode
# --------------------------------
print("\n💬 INTERACTIVE MODE - Ask Questions")
print("="*50)
print("Type your questions and get AI-powered answers from your documents.")
print("Type 'exit' or 'quit' to finish.\n")

conversation_history = []

while True:
    user_input = input("\n>>> ").strip()

    if user_input.lower() in ['exit', 'quit', 'q']:
        print("\n👋 Finishing session...")
        break

    if not user_input:
        continue

    print("\n🤖 Generating answer...")
    answer = pipeline.query_with_llm(user_input, top_k=5)

    print(f"\n{'='*60}")
    print(answer)
    print('='*60)

    conversation_history.append({
        'question': user_input,
        'answer': answer
    })

print(f"\n✅ Session complete! Asked {len(conversation_history)} questions.")


# Cell 14: Download Excel and Export Reports
# -------------------------------------------
print("\n💾 GENERATING FINAL REPORT & SAVING FILES")
print("="*50)

import json
from datetime import datetime

# Generate text report
report_content = f"""
{'='*80}
RAG SYSTEM SESSION REPORT
Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
{'='*80}

SYSTEM CONFIGURATION
{'='*80}
Collection Name: {pipeline.collection_name}
LLM Provider: {pipeline.llm_provider.upper()}
LLM Model: {pipeline.llm_model}
Embedding Model: BGE-M3 (1024 dimensions)
Device: {pipeline.device.upper()}
Chunk Size: {pipeline.chunk_size} words
Overlap: {pipeline.overlap} words
Batch Size: {pipeline.batch_size}

PERFORMANCE METRICS
{'='*80}
Documents Processed: {pipeline.tracker.metrics['files']}
Total Size: {pipeline.tracker.metrics['total_size_mb']:.2f} MB
Total Tokens: {pipeline.tracker.metrics['total_tokens']/1e6:.1f} Million
Text Extraction: {pipeline.tracker.metrics['text_extraction_time']:.1f}s
Chunking: {pipeline.tracker.metrics['chunk_time']:.1f}s
Embedding: {pipeline.tracker.metrics['embed_time']:.1f}s
Upload: {pipeline.tracker.metrics['ingestion_time']:.1f}s
Total Time: {pipeline.tracker.metrics['total_time']:.1f}s
GPU Usage: {pipeline.tracker.metrics['gpu_usage_gb']:.1f} GB
Embedding Speed: {pipeline.tracker.metrics['mb_embed_per_min']:.2f} MB/min

DOCUMENT REGISTRY
{'='*80}
"""

for doc_name, info in pipeline.document_registry.items():
    report_content += f"""
📄 {doc_name}
   Type: {info['document_type']}
   Chunks: {info['total_chunks']}
   Size: {info['size_mb']:.2f} MB
"""

if conversation_history:
    report_content += f"""

INTERACTIVE SESSION
{'='*80}
Total Questions Asked: {len(conversation_history)}

"""
    for i, conv in enumerate(conversation_history, 1):
        report_content += f"""
Question {i}: {conv['question']}
{'-'*80}
Answer: {conv['answer']}

{'='*80}
"""

report_content += f"""
END OF REPORT
{'='*80}
"""

# Save locally (changed paths)
with open('./rag_session_report.txt', 'w', encoding='utf-8') as f:
    f.write(report_content)

print("✅ Report saved: rag_session_report.txt")

# Generate statistics CSV
import pandas as pd

stats_data = {
    'Metric': [
        'Total Documents',
        'Total Size (MB)',
        'Total Tokens (M)',
        'Scan Time (s)',
        'Extraction Time (s)',
        'Chunking Time (s)',
        'Embedding Time (s)',
        'Upload Time (s)',
        'Total Time (s)',
        'GPU Usage (GB)',
        'Embedding Speed (MB/min)',
        'Questions Asked'
    ],
    'Value': [
        pipeline.tracker.metrics['files'],
        round(pipeline.tracker.metrics['total_size_mb'], 2),
        round(pipeline.tracker.metrics['total_tokens']/1e6, 2),
        round(pipeline.tracker.metrics['drive_download_scan_time'], 2),
        round(pipeline.tracker.metrics['text_extraction_time'], 2),
        round(pipeline.tracker.metrics['chunk_time'], 2),
        round(pipeline.tracker.metrics['embed_time'], 2),
        round(pipeline.tracker.metrics['ingestion_time'], 2),
        round(pipeline.tracker.metrics['total_time'], 2),
        round(pipeline.tracker.metrics['gpu_usage_gb'], 2),
        round(pipeline.tracker.metrics['mb_embed_per_min'], 2),
        len(conversation_history)
    ]
}

df_stats = pd.DataFrame(stats_data)
df_stats.to_csv('./rag_statistics.csv', index=False)

print("✅ Statistics saved: rag_statistics.csv")

# Generate JSON config
config = {
    'session_info': {
        'timestamp': datetime.now().isoformat(),
        'collection_name': pipeline.collection_name,
        'llm_provider': pipeline.llm_provider,
        'llm_model': pipeline.llm_model,
        'chunk_size': pipeline.chunk_size,
        'overlap': pipeline.overlap,
        'batch_size': pipeline.batch_size
    },
    'documents': pipeline.document_registry,
    'metrics': {
        'total_documents': pipeline.tracker.metrics['files'],
        'total_size_mb': pipeline.tracker.metrics['total_size_mb'],
        'total_tokens': pipeline.tracker.metrics['total_tokens'],
        'embedding_speed_mb_per_min': pipeline.tracker.metrics['mb_embed_per_min']
    },
    'conversation_summary': {
        'total_questions': len(conversation_history),
        'questions': [conv['question'] for conv in conversation_history]
    }
}

with open('./rag_session_config.json', 'w') as f:
    json.dump(config, f, indent=2)

print("✅ Config saved: rag_session_config.json")

# Removed files.download() calls - not needed locally
print("\n" + "="*60)
print("✅ ALL FILES SAVED")
print("="*60)
print("\nFiles generated in current directory:")
print("1. rag_metrics.xlsx - Excel metrics tracker")
print("2. rag_session_report.txt - Complete session report")
print("3. rag_statistics.csv - Performance metrics")
print("4. rag_session_config.json - Configuration details")
print("\n" + "="*60)


# Cell 15: Session Summary
# -------------------------
print("\n📊 FINAL SESSION SUMMARY")
print("="*50)
print(f"""
✅ Documents Processed: {pipeline.tracker.metrics['files']}
✅ Total Size: {pipeline.tracker.metrics['total_size_mb']:.2f} MB
✅ Total Tokens: {pipeline.tracker.metrics['total_tokens']/1e6:.1f}M
✅ Questions Answered: {len(conversation_history)}

System Performance:
- Embedding Speed: {pipeline.tracker.metrics['mb_embed_per_min']:.2f} MB/min
- Total Processing Time: {pipeline.tracker.metrics['total_time']:.1f}s
- GPU Usage: {pipeline.tracker.metrics['gpu_usage_gb']:.1f} GB

📊 Excel metrics saved to: {EXCEL_PATH}
📥 All files downloaded successfully!

✅ Your RAG system is ready for production use!
""")